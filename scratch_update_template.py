import openpyxl
import os

TEMPLATE_DIR = os.path.join('app', 'templates', 'excel', 'OrdenProduccion')
INPUT_PATH = os.path.join(TEMPLATE_DIR, 'Book1.xlsx')
OUTPUT_PATH = os.path.join(TEMPLATE_DIR, 'Book2.xlsx')

def main():
    print(f"Loading {INPUT_PATH}...")
    wb = openpyxl.load_workbook(INPUT_PATH)
    ws = wb['IMPRIMIR OP']

    # Insert 4 rows at row 8 (this shifts row 8 down to 12)
    print("Inserting 4 rows at index 8...")
    ws.insert_rows(8, 4)

    for row in range(8, 12):
        ws.row_dimensions[row].height = 15

    # Update the shifted labels (were at row 8-10, now at 12-14)
    print("Updating labels in shifted rows...")
    ws['F12'] = 'Peso Neto Golpe:'
    ws['F13'] = 'Peso Colada:'
    ws['F14'] = 'Peso Tiro:'
    ws['F15'] = 'Merma:'
    ws['F16'] = 'Fecha:'
    
    # Let's also add the composition headers at row 8 just to be nice
    ws['B8'] = 'Composición'
    ws['C8'] = 'Cav'
    ws['D8'] = 'Peso Unit'
    ws['E8'] = 'Subtotal'

    # Clear any existing page breaks
    print("Clearing page breaks...")
    ws.row_breaks = openpyxl.worksheet.pagebreak.RowBreak()

    print(f"Saving to {OUTPUT_PATH}...")
    wb.save(OUTPUT_PATH)
    print("Done!")

if __name__ == '__main__':
    main()
