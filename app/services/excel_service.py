"""
Servicio de generación de Excel para Órdenes de Producción.
Llena la plantilla "IMPRIMIR OP" con los datos de una orden.
"""
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime
import os

# Ruta a la plantilla Excel
TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), '..', 'templates', 'excel', 'OrdenProduccion', 'Book2.xlsx')


def generar_op_excel(orden) -> BytesIO:
    """
    Genera un archivo Excel llenando la pestaña "IMPRIMIR OP" con los datos de la orden.
    
    Args:
        orden: Objeto OrdenProduccion con sus lotes cargados
        
    Returns:
        BytesIO: Buffer con el archivo Excel listo para descarga
    """
    # Cargar plantilla
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb['IMPRIMIR OP']
    
    # Obtener datos calculados
    resumen = orden.resumen_totales
    
    # =========================================================================
    # 1. CABECERA (Filas 1-12)
    # =========================================================================
    ws['D1'] = orden.numero_op
    ws['C3'] = orden.fecha_creacion.strftime('%Y-%m-%d') if orden.fecha_creacion else ''
    ws['C4'] = orden.producto or ''
    ws['G4'] = orden.maquina_id or ''
    ws['C5'] = resumen.get('Cantidad DOC', 0)
    ws['B7'] = orden.molde or ''
    
    # === COMPOSICION DEL MOLDE (Filas 8-11, nuevas) ===
    # Llenar hasta 4 piezas
    composicion = orden.snapshot_composicion
    for i, snap in enumerate(composicion[:4]):
        row = 8 + i
        ws[f'B{row}'] = snap.pieza.piezas if snap.pieza else f'Pieza {i+1}'
        ws[f'C{row}'] = snap.cavidades
        ws[f'D{row}'] = snap.peso_unit_gr
        ws[f'E{row}'] = snap.peso_subtotal_gr
    
    # === PARAMETROS TECNICOS (Filas 12-16, desplazadas) ===
    ws['C12'] = resumen.get('Días', 0)
    ws['G12'] = orden.calculo_peso_neto_golpe or 0
    ws['C13'] = orden.snapshot_horas_turno or 24
    ws['G13'] = orden.snapshot_peso_colada_gr or 0   # Peso Colada
    ws['C14'] = orden.calculo_cavidades_totales or 1
    ws['G14'] = orden.calculo_peso_tiro_gr or 0      # Peso Tiro
    
    # Coladas por hora: 3600 / tiempo_ciclo
    coladas_hora = 0
    if orden.snapshot_tiempo_ciclo and orden.snapshot_tiempo_ciclo > 0:
        coladas_hora = 3600 / orden.snapshot_tiempo_ciclo
    ws['C15'] = coladas_hora
    ws['G15'] = resumen.get('%Merma', 0)             # Merma
    
    # Rango de fechas
    fecha_inicio = orden.fecha_inicio.strftime('%d/%m') if orden.fecha_inicio else ''
    fecha_fin_str = resumen.get('F. Fin', '')
    if fecha_fin_str:
        try:
            fecha_fin = datetime.fromisoformat(fecha_fin_str)
            ws['G16'] = f"{fecha_inicio} a {fecha_fin.strftime('%d/%m')}"
        except:
            ws['G16'] = fecha_inicio
    else:
        ws['G16'] = fecha_inicio
    
    # =========================================================================
    # 2. TABLA DE COLORES / PRODUCCIÓN (Filas 18-25, desplazadas +4)
    # =========================================================================
    lotes = orden.lotes[:6]  # Máximo 6 colores
    
    # Totales para sumar
    total_peso = 0
    total_merma = 0
    total_coladas = 0
    
    for i, lote in enumerate(lotes):
        row = 19 + i  # Filas 19-24
        lote_data = lote.to_dict()
        
        # Nombre del color
        ws[f'B{row}'] = lote_data.get('Color', 'Sin Color')
        
        # Peso producción por color (peso base + extra)
        peso_lote = lote_data.get('TOTAL + EXTRA (Kg)', 0)
        ws[f'C{row}'] = peso_lote
        total_peso += peso_lote
        
        # Merma a recuperar (calculada)
        merma_pct = resumen.get('%Merma', 0)
        merma_lote = peso_lote * merma_pct if merma_pct else 0
        ws[f'E{row}'] = merma_lote
        total_merma += merma_lote
        
        # Cantidad coladas
        coladas = lote_data.get('coladas_calculadas', 0)
        ws[f'G{row}'] = coladas
        total_coladas += coladas
    
    # Fila de totales (25)
    ws['C25'] = total_peso
    ws['E25'] = total_merma
    ws['G25'] = total_coladas
    
    # =========================================================================
    # 3. MATERIA PRIMA TOTALES (Filas 27-32, desplazadas +4)
    # =========================================================================
    # Acumular materiales por tipo
    materiales_totales = {'VIRGEN': 0, 'VIRGEN_2': 0, 'MOLIDO': 0}
    materiales_nombres = {'VIRGEN': '', 'VIRGEN_2': '', 'MOLIDO': ''}
    
    for lote in lotes:
        for mat in lote.materias_primas:
            tipo = mat.materia.tipo if mat.materia else 'VIRGEN'
            peso = mat.peso_kg
            materiales_totales[tipo] = materiales_totales.get(tipo, 0) + peso
            if not materiales_nombres.get(tipo) and mat.materia:
                materiales_nombres[tipo] = mat.materia.nombre
    
    ws['D29'] = materiales_nombres.get('VIRGEN', '')
    ws['F29'] = materiales_totales.get('VIRGEN', 0)
    ws['D30'] = materiales_nombres.get('VIRGEN_2', '')
    ws['F30'] = materiales_totales.get('VIRGEN_2', 0)
    ws['D31'] = materiales_nombres.get('MOLIDO', '')
    ws['F31'] = materiales_totales.get('MOLIDO', 0)
    
    total_material = sum(materiales_totales.values())
    ws['F32'] = total_material
    
    # =========================================================================
    # 4. MATERIA PRIMA POR COLOR (Filas 34-42, desplazadas +4)
    # =========================================================================
    for i, lote in enumerate(lotes):
        row = 36 + i  # Filas 36-41
        
        lote_data = lote.to_dict()
        ws[f'B{row}'] = lote_data.get('Color', 'Sin Color')
        
        # Procesar materiales del lote
        materiales_lote = {'VIRGEN': None, 'VIRGEN_2': None, 'MOLIDO': None}
        for mat in lote.materias_primas:
            tipo = mat.materia.tipo if mat.materia else 'VIRGEN'
            nombre = mat.materia.nombre if mat.materia else ''
            fraccion = mat.fraccion
            peso = mat.peso_kg
            materiales_lote[tipo] = {
                'nombre': f"{nombre} = {int(fraccion*6)}/6" if fraccion else nombre,
                'peso': peso
            }
        
        # Virgen 1 (C, D)
        if materiales_lote['VIRGEN']:
            ws[f'C{row}'] = materiales_lote['VIRGEN']['nombre']
            ws[f'D{row}'] = materiales_lote['VIRGEN']['peso']
        
        # Virgen 2 (E, F)
        if materiales_lote['VIRGEN_2']:
            ws[f'E{row}'] = materiales_lote['VIRGEN_2']['nombre']
            ws[f'F{row}'] = materiales_lote['VIRGEN_2']['peso']
        else:
            ws[f'F{row}'] = 0
        
        # Segunda (G, H)
        if materiales_lote['MOLIDO']:
            ws[f'G{row}'] = materiales_lote['MOLIDO']['nombre']
            ws[f'H{row}'] = materiales_lote['MOLIDO']['peso']
    
    # Totales materiales por color (fila 42)
    ws['D42'] = materiales_totales.get('VIRGEN', 0)
    ws['F42'] = materiales_totales.get('VIRGEN_2', 0)
    ws['H42'] = materiales_totales.get('MOLIDO', 0)
    
    # =========================================================================
    # 5. COLORANTES (Filas 44-60, desplazadas +4)
    # =========================================================================
    # Grupo 1: Colores 0-2 (filas 45-52)
    if len(lotes) >= 1:
        _llenar_colorantes_grupo(ws, lotes[0:3], start_header_row=45, start_data_row=46)
    
    # Grupo 2: Colores 3-5 (filas 53-60)
    if len(lotes) >= 4:
        _llenar_colorantes_grupo(ws, lotes[3:6], start_header_row=53, start_data_row=54)
    
    # =========================================================================
    # 6. QR CODE (Celda C65 - bajado 4 casillas desde C61)
    # =========================================================================
    from app.services.qr_service import generar_qr_imagen
    
    try:
        qr_buffer = generar_qr_imagen(orden, size=240)  # QR más grande
        qr_img = XLImage(qr_buffer)
        qr_img.width = 200
        qr_img.height = 200
        ws.add_image(qr_img, 'C65')
    except Exception as e:
        # Si falla el QR, continuamos sin él
        print(f"Error generando QR: {e}")
    
    # =========================================================================
    # 7. CONFIGURAR ZONA DE IMPRESIÓN Y SALTOS DE PÁGINA
    # =========================================================================
    from openpyxl.worksheet.pagebreak import Break
    
    # Limpiar saltos viejos y forzar ajuste al ancho
    ws.row_breaks = openpyxl.worksheet.pagebreak.RowBreak()
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    # Insertar salto de página justo antes de colorantes (Fila 44)
    # openpyxl inserta el salto *después* de la fila id, así que usamos id=43
    ws.row_breaks.append(Break(id=43))
    
    # Definir area de impresión hasta el final del QR (Fila 75 aprox)
    ws.print_area = 'A1:H75'

    # =========================================================================
    # GUARDAR A BUFFER
    # =========================================================================
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def _llenar_colorantes_grupo(ws, lotes_grupo, start_header_row, start_data_row):
    """
    Llena un grupo de colorantes (3 colores) en el Excel.
    
    Args:
        ws: Hoja de Excel
        lotes_grupo: Lista de hasta 3 lotes
        start_header_row: Fila donde van los nombres de colores
        start_data_row: Fila donde empiezan los colorantes
    """
    # Columnas para cada color del grupo
    columnas = [('C', 'D'), ('E', 'F'), ('G', 'H')]
    
    for idx, lote in enumerate(lotes_grupo):
        if idx >= 3:
            break
        
        col_nombre, col_gramos = columnas[idx]
        
        # Header: Nombre del color
        ws[f'{col_nombre}{start_header_row}'] = lote.color_produccion_rel.nombre if lote.color_produccion_rel else 'Sin Color'
        ws[f'{col_gramos}{start_header_row}'] = 'Gr.'
        
        # Colorantes (hasta 7 filas)
        # Data: Colorantes
        for i, pig in enumerate(lote.colorantes):
            row = start_data_row + i
            if row > start_data_row + 20: # Limite seguridad
                break
                
            ws[f'{col_nombre}{row}'] = pig.pigmento.nombre if pig.pigmento else ''
            ws[f'{col_gramos}{row}'] = (pig.gramos or 0) * 2  # Multiplicador x2 por requerimiento de impresión

