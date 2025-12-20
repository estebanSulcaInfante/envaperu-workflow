import openpyxl

def inspect_formulas():
    path = 'app/templates/excel/OP1322-BALDE ROMANO.xlsm'
    wb = openpyxl.load_workbook(path, data_only=False)
    ws = wb['IMPRIMIR OP']
    
    print("--- Fórmulas en IMPRIMIR OP ---")
    
    # Filas de ejemplo (15, 16, 17) donde suelen estar los lotes
    for row in range(15, 18):
        print(f"\nFila {row}:")
        print(f"  C{row} (Peso): {ws[f'C{row}'].value}")
        print(f"  E{row} (Merma Rec.): {ws[f'E{row}'].value}")
        print(f"  G{row} (Coladas): {ws[f'G{row}'].value}")

    print("\n--- Totales Materia Prima ---")
    print(f"F25 (Virgen): {ws['F25'].value}")
    print(f"F26 (Virgen 2): {ws['F26'].value}")
    print(f"F27 (Molido): {ws['F27'].value}")
    print(f"F28 (TOTAL): {ws['F28'].value}")

    print("\n--- Valores Calculados (Data Only) ---")
    wb_data = openpyxl.load_workbook(path, data_only=True)
    ws_data = wb_data['IMPRIMIR OP']
    
    for row in range(15, 18):
        print(f"\nFila {row}:")
        print(f"  C{row} (Peso): {ws_data[f'C{row}'].value}")
        print(f"  E{row} (Merma Rec.): {ws_data[f'E{row}'].value}")
        print(f"  G{row} (Coladas): {ws_data[f'G{row}'].value}")

    print("\n--- Valores Totales Materia Prima ---")
    print(f"F25 (Virgen): {ws_data['F25'].value}")
    print(f"F26 (Virgen 2): {ws_data['F26'].value}")
    print(f"F27 (Molido): {ws_data['F27'].value}")
    print(f"F28 (TOTAL): {ws_data['F28'].value}")

if __name__ == "__main__":
    inspect_formulas()
