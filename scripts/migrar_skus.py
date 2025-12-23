import openpyxl
import os
import sys

# Agregar root al path para imports
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from app import create_app
from app.extensions import db
from app.models.producto import ProductoTerminado, Pieza, ProductoPieza

def safe_float(val):
    if not val:
        return None
    s = str(val).strip()
    if s.startswith('#'): # Errores excel #DIV/0!, #N/A, etc
        return None
    try:
        return float(val)
    except:
        return None

def migrar_productos_terminados(ws):
    print("Migrando Productos Terminados...")
    count = 0
    # Iterar desde fila 2
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[8]: # Si no hay SKU PT, saltar (fila vacia o invalida)
            continue
            
        sku = str(row[8]).strip()
        # Clean sku if formula error (though unlikely as PK)
        if sku.startswith('#'): continue

        # Verificar si existe
        pt = db.session.get(ProductoTerminado, sku)
        if not pt:
            pt = ProductoTerminado(cod_sku_pt=sku)
            
        pt.cod_linea_num = row[0]
        pt.linea = row[1]
        pt.cod_familia = row[2]
        pt.familia = row[3]
        pt.cod_producto = row[4]
        pt.producto = str(row[5]).strip() if row[5] else None 
        pt.cod_color = row[6]
        pt.familia_color = row[7]
        # row[8] es SKU
        pt.um = row[9]
        pt.doc_x_paq = row[10]
        pt.doc_x_bulto = row[11]
        
        # Floats seguros
        pt.peso_g = safe_float(row[12])
        pt.precio_estimado = safe_float(row[13])
        pt.precio_sin_igv = safe_float(row[14])
        pt.indicador_x_kg = safe_float(row[15])
        
        pt.status = row[16]
        pt.codigo_barra = str(row[17]) if row[17] else None
        pt.marca = row[18]
        pt.nombre_gs1 = row[19]
        pt.obs = row[20]
        
        db.session.add(pt)
        count += 1
        
    db.session.commit()
    print(f"--> {count} Productos Terminados procesados.")

def migrar_piezas(ws):
    print("Migrando Piezas...")
    count = 0
    
    # Pre-cargar mapa de Productos {Nombre: SKU} para evitar queries n+1
    mapa_productos = {}
    pts = ProductoTerminado.query.with_entities(ProductoTerminado.producto, ProductoTerminado.cod_sku_pt).all()
    for nombre, sku in pts:
        if nombre:
            mapa_productos[nombre.upper()] = sku # Normalizar mayusculas
            
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]: # Si no hay SKU
            continue
            
        sku_pieza = str(row[0]).strip()
        
        # Crear o Actualizar
        pieza = db.session.get(Pieza, sku_pieza)
        if not pieza:
            pieza = Pieza(sku=sku_pieza)
            
        pieza.cod_linea = row[1]
        pieza.linea = row[2]
        pieza.familia = row[3]
        
        nombre_producto_padre = str(row[4]).strip() if row[4] else None
        # Guardar nombre para crear relacion despues
        _nombre_producto_map[sku_pieza] = nombre_producto_padre
        
        pieza.cod_pieza = row[5]
        pieza.piezas = row[6]
        pieza.cod_col = row[7]
        pieza.tipo_color = row[8]
        pieza.cavidad = row[9]
        pieza.peso = safe_float(row[10])
        pieza.cod_extru = row[11]
        pieza.tipo_extruccion = row[12]
        pieza.cod_mp = row[13]
        pieza.mp = row[14]
        pieza.cod_color = row[15]
        pieza.color = row[16]
                
        db.session.add(pieza)
        count += 1
        
    db.session.commit()
    print(f"--> {count} Piezas procesadas.")
    return _nombre_producto_map

# Mapa temporal para guardar relaciones
_nombre_producto_map = {}

def crear_relaciones_producto_pieza(mapa_nombres):
    """Crea registros en la tabla intermedia ProductoPieza."""
    print("Creando relaciones Producto-Pieza...")
    count = 0
    
    # Cargar mapa de productos por nombre
    mapa_productos = {}
    pts = ProductoTerminado.query.with_entities(ProductoTerminado.producto, ProductoTerminado.cod_sku_pt).all()
    for nombre, sku in pts:
        if nombre:
            mapa_productos[nombre.upper()] = sku
    
    for pieza_sku, nombre_producto in mapa_nombres.items():
        if not nombre_producto:
            continue
            
        sku_producto = mapa_productos.get(nombre_producto.upper())
        if sku_producto:
            # Verificar si ya existe la relacion
            existe = ProductoPieza.query.filter_by(
                producto_terminado_id=sku_producto, 
                pieza_sku=pieza_sku
            ).first()
            
            if not existe:
                rel = ProductoPieza(
                    producto_terminado_id=sku_producto,
                    pieza_sku=pieza_sku,
                    cantidad=1
                )
                db.session.add(rel)
                count += 1
    
    db.session.commit()
    print(f"--> {count} relaciones creadas.")

def main():
    app = create_app()
    with app.app_context():
        # Crear tablas si no existen
        db.create_all()
        
        base_dir = r"c:\Users\esteb\envaperu-workflow\app\templates\excel\Skus"
        
        # 1. Productos Terminados (Padres)
        file_pt = os.path.join(base_dir, "SKU PRODUCTOS TERMINADOS 2025.xlsx")
        wb_pt = openpyxl.load_workbook(file_pt, data_only=True)
        migrar_productos_terminados(wb_pt.active)
        
        # 2. Piezas
        file_piezas = os.path.join(base_dir, "SKU PIEZAS 2025.xlsx")
        wb_p = openpyxl.load_workbook(file_piezas, data_only=True)
        
        # Mapa acumulado para relaciones
        mapa_relaciones = {}
        
        for sheet_name in wb_p.sheetnames:
            if "SKU PIEZAS" in sheet_name:
                print(f"Procesando hoja piezas: {sheet_name}")
                mapa_partial = migrar_piezas(wb_p[sheet_name])
                mapa_relaciones.update(mapa_partial)
        
        # 3. Crear relaciones en tabla intermedia
        crear_relaciones_producto_pieza(mapa_relaciones)

if __name__ == "__main__":
    main()

