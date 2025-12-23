import openpyxl
import os

def analizar_excel(path, nombre_tipo):
    print(f"\n{'='*50}")
    print(f"ANALIZANDO: {nombre_tipo}")
    print(f"Archivo: {path}")
    print(f"{'='*50}")

    if not os.path.exists(path):
        print(f"ERROR: No se encuentra el archivo {path}")
        return

    try:
        # Cargar CON fórmulas
        wb = openpyxl.load_workbook(path, data_only=False)
        print(f"Hojas encontradas: {wb.sheetnames}")
        
        ws = wb.active
        print(f"Hoja Activa: '{ws.title}'")
        
        # Leer cabeceras
        headers = [cell.value for cell in ws[1]]
        
        print("\n[CABECERAS]")
        for idx, h in enumerate(headers):
            print(f"  Col {idx+1}: {h}")

        print("\n[FÓRMULAS/VALORES (Fila 2)]")
        # Leer fila 2 para ver fórmulas
        row_2 = list(ws[2]) 
        for idx, cell in enumerate(row_2):
            val = cell.value
            header = headers[idx] if idx < len(headers) else f"Col {idx+1}"
            # Si empieza con =, es formula
            es_formula = str(val).startswith('=') if val else False
            print(f"  {header}: {val} {'[FORMULA]' if es_formula else ''}")


    except Exception as e:
        print(f"Error analizando {path}: {e}")

def main():
    base_dir = r"c:\Users\esteb\envaperu-workflow\app\templates\excel\Skus"
    
    file_piezas = os.path.join(base_dir, "SKU PIEZAS 2025.xlsx")
    analizar_excel(file_piezas, "SKU PIEZAS")
    
    file_productos = os.path.join(base_dir, "SKU PRODUCTOS TERMINADOS 2025.xlsx")
    analizar_excel(file_productos, "SKU PRODUCTOS TERMINADOS")

if __name__ == "__main__":
    main()
