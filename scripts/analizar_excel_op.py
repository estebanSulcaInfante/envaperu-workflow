"""
Script para analizar la estructura del archivo Excel de Órdenes de Producción
y crear un mapping de las celdas de la pestaña "IMPRIMIR OP"
"""
import openpyxl
import warnings
import json
from openpyxl.utils import get_column_letter

warnings.filterwarnings('ignore')

EXCEL_PATH = 'app/templates/excel/OP1322-BALDE ROMANO.xlsm'

def analizar_hoja(ws, nombre_hoja, max_rows=100, max_cols=20):
    print(f"\n{'='*60}")
    print(f"HOJA: {nombre_hoja}")
    print(f"{'='*60}")
    print(f"Rango: {ws.dimensions}")
    print(f"Max filas: {ws.max_row}, Max columnas: {ws.max_column}")
    
    print("\n--- CELDAS CON CONTENIDO ---")
    celdas = []
    for row in range(1, min(max_rows + 1, ws.max_row + 1)):
        for col in range(1, min(max_cols + 1, ws.max_column + 1)):
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                col_letter = get_column_letter(col)
                valor = str(cell.value)[:100]
                celdas.append({
                    'celda': f'{col_letter}{row}',
                    'fila': row,
                    'columna': col,
                    'col_letra': col_letter,
                    'valor': valor
                })
                print(f"  {col_letter}{row}: {valor}")
    
    return celdas

def main():
    print("Cargando archivo Excel...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    
    print(f"\nPestañas disponibles: {wb.sheetnames}")
    
    # Analizar la pestaña IMPRIMIR OP
    if 'IMPRIMIR OP' in wb.sheetnames:
        ws = wb['IMPRIMIR OP']
        celdas = analizar_hoja(ws, 'IMPRIMIR OP', max_rows=80, max_cols=15)
        
        # Guardar mapping en JSON
        with open('scripts/excel_mapping_imprimir_op.json', 'w', encoding='utf-8') as f:
            json.dump(celdas, f, ensure_ascii=False, indent=2)
        print(f"\n✅ Mapping guardado en scripts/excel_mapping_imprimir_op.json")
    
    # También ver la pestaña Ordenes para entender las fórmulas
    if 'Ordenes' in wb.sheetnames:
        ws_ordenes = wb['Ordenes']
        analizar_hoja(ws_ordenes, 'Ordenes', max_rows=30, max_cols=30)

if __name__ == '__main__':
    main()
